import json
import os
import pandas as pd
import sqlite3
import time

from slpp import slpp as lua
from openpyxl import load_workbook
from openpyxl.styles import Font

from settings import *


class AccData:
    def __init__(self, path):
        def table_to_dict(path, string_to_remove):
            """Decodes lua SavedVariables table as dict."""
            if not os.path.exists(path):
                return {}
            data = {}
            with open(path, encoding="utf-8") as file:
                if file:
                    data = lua.decode(file.read().replace(string_to_remove, ''))
                    return data

        data = {} # container for saved_var dicts 
        for name, info in LUA_FILES.items():
            file_path = path + '\\' + info['file_name']
            data[name] = table_to_dict(file_path, info['string_to_remove'])

        self.path = path
        self.characters = data['Data']['charData']
        self.realms = data['Data']['realmData']
        self.accounting = data['Accounting']


class Character:
    def __init__(self, name, realm, account):
        self.name = name
        self.realm = realm
        self.account = account
        self.name_realm_with_spaces = '-'.join((self.name, self.realm))
        self.name_realm = self.name_realm_with_spaces.replace(' ', '')


class Farmer(Character):
    def __init__(self, data):
        super().__init__(data[0], data[1], data[2])
        self.char_class = data[3]
        self.intro_completed = None
        self.max_riding = None
        self.professions = {}
        self.recipe_ranks = {}

    def update_info(self, data):
        self.intro_completed = data.get('introCompleted', None)
        self.max_riding = data.get('maxRidingSkill', None)
        self.professions = {}
        for profession_id, rating in data.get('professions', None).items():
            if type(rating) is int:
                self.professions[profession_id] = rating
        self.recipe_ranks = {}
        for recipe, rank in data.get('recipeRanks', None).items():
            self.recipe_ranks[recipe] = rank


class Banker(Character):
    def __init__(self, data):
        super().__init__(data[0], data[1], data[2])
        self.bank_number = 0 if data[3] == 'deposit' else int(data[3].replace('gbank', ''))
        self.bank_gold = None

    def update_info(self, money):
        self.bank_gold = money // 10000 if money else None


class RealmData:
    def __init__(self, data):
        self.name = data[0]
        self.slug = data[1]
        self.code = data[2]
        self.last_update = data[3]
        self.seller_name = data[4]
        self.seller_name_realm = '-'.join((self.seller_name, self.name.replace(' ', '')))
        self.seller_name_realm_with_spaces = '-'.join((self.seller_name, self.name))

        self.inventory = {}

    def update_inventory(self, auctions, inventory):
        for item_id, qty in inventory.items():
            self.inventory[item_id] = self.inventory.get(item_id, {})
            self.inventory[item_id]['bags'] = self.inventory[item_id].get('bags', 0) + qty
        for item_id, qty in auctions.items():
            self.inventory[item_id] = self.inventory.get(item_id, {})
            self.inventory[item_id]['ah'] = self.inventory[item_id].get('ah', 0) + qty


class WowData:
    def __init__(self):
        def farmer_objects_dict():
            conn = sqlite3.connect(CHAR_DB)
            c = conn.cursor()
            c.execute("""SELECT name, realm, account, class FROM char_db
                    WHERE role=? AND (type!=? or type IS NULL)""",
                    ('farmer', 'inactive'))
            rows = c.fetchall()
            conn.close()

            farmers = {}
            for row in rows:
                name_realm = '-'.join((row[0], row[1]))
                farmers[name_realm] = Farmer(row)
            return farmers

        def banker_objects_dict():
            conn = sqlite3.connect(CHAR_DB)
            c = conn.cursor()
            c.execute("""SELECT name, realm, account, type FROM char_db
                      WHERE role=?""",
                      ('banker', ))
            rows = c.fetchall()
            conn.close()

            bankers = {}
            for row in rows:
                banker = Banker(row)
                bankers[banker.name_realm_with_spaces] = banker
            return bankers

        def account_objects_dict():
            accounts = {}
            for acc_number, path in LUA_PATHS.items():
                accounts[int(acc_number)] = AccData(path)
            return accounts  

        def create_output_db():
            """write doc"""
            
            conn = sqlite3.connect(FARMERS_DB)
            c = conn.cursor()
            c.execute("""CREATE TABLE IF NOT EXISTS farmers (
                name TEXT NOT NULL,
                realm TEXT NOT NULL,
                account INTEGER NOT NULL,
                class TEXT,
                intro_complete INTEGER,
                Herbalism INTEGER,
                Mining INTEGER,
                "Zin'anthid" INTEGER,
                "Osmenite Deposit" INTEGER,
                "Osmenite Seam" INTEGER,
                max_riding INTEGER)""")
            conn.close()

        def realms_inventory():
            realms = {}
            conn = sqlite3.connect(REALMS)
            c = conn.cursor()
            c.execute("SELECT name, slug, code, last_update, seller FROM realms")
            for data in c.fetchall():
                realms[data[0]] = RealmData(data)

            conn.close()
            return realms

        self.farmers = farmer_objects_dict()
        self.bankers = banker_objects_dict()
        self.accounts = account_objects_dict()
        self.realms = realms_inventory()
        create_output_db()

    def update_farmers(self):
        """write doc"""
        for acc_num in self.accounts.keys():
            char_list = self.accounts[acc_num].characters
            for full_name, char_data in char_list.items():
                self.farmers[full_name].update_info(char_data)

    def write_farmers_db(self):
        """write doc"""
        conn = sqlite3.connect(FARMERS_DB)
        c = conn.cursor()
        c.execute("DELETE FROM farmers")

        for full_name, farmer in self.farmers.items():
            values = (farmer.name,
                      farmer.realm,
                      farmer.account,
                      farmer.char_class,
                      1 if farmer.intro_completed else 0,
                      farmer.professions.get(2549, None),
                      farmer.professions.get(2565, None),
                      farmer.recipe_ranks.get("Zin'anthid", None),
                      farmer.recipe_ranks.get('Osmenite Deposit', None),
                      farmer.recipe_ranks.get('Osmenite Seam', None),
                      1 if farmer.max_riding else 0,)
            c.execute("""INSERT INTO farmers
                    (name, realm, account, class, intro_complete, Herbalism, Mining,
                    "Zin'anthid", "Osmenite Deposit", "Osmenite Seam", max_riding)
                    VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    values)
        conn.commit()
        conn.close()

    def write_excel_table(self):
        wb = load_workbook(EXCEL_PATH, read_only=False, keep_vba=True)
        ws = wb["Farming"]
        selection = ws["B32:K57"]

        conn = sqlite3.connect(FARMERS_DB)
        c = conn.cursor()

        realm_list = ['Kazzak', 'Twisting Nether', 'Silvermoon', 'Draenor', 'Tarren Mill', 'Ragnaros', 'Ravencrest', 'Argent Dawn', 'Sylvanas', 'Frostmane', 'Burning Blade', 'Blackmoore', 'Blackrock', 'Blackhand', 'Antonidas', 'Hyjal', 'Archimonde', 'Thrall', 'Eredar', 'Ysondre', 'Dalaran', 'Onyxia', 'Nefarian', 'The Maelstrom', 'Frostwolf', 'Aegwynn']
        realm_codes = ['KZ','TN','SM','DR','TM','RG','RV','AD','SV','FM','BB','BM','BR','BH','AN','HJ','AR','TH','ER','YS','DL','ON','NF','MA','FW','AE']
        row_offset = 32
        col_offset = 2

        for row in selection:
            for cell in row:
                c.execute("SELECT * FROM farmers WHERE realm=? AND account=?",
                        (realm_list[cell.row - row_offset], cell.col_idx - col_offset))
                db_row = c.fetchone()
                if db_row:
                    cell.value = realm_codes[cell.row - row_offset]
                    # char info
                    char_class = db_row[3]
                    intro_complete = True if db_row[4] == 1 else False
                    herbalism_rating = db_row[5] or 0
                    mining_rating = db_row[6] or 0
                    zin = db_row[7] or 0
                    os_deposit = db_row[8] or 0
                    os_seam = db_row[9] or 0
                    max_riding = True if db_row[10] == 1 else False

                    cell_bold = False if char_class == 'dh' else True
                    cell_underline = 'single' if char_class == 'dh' else 'none'
                    cell_color = "000000"

                    if intro_complete and not max_riding:
                        cell_color = "ff26ff"
                    elif zin == 3 and os_deposit == 3 and os_seam == 3:
                        cell_color = "14ccf5"
                    elif zin == 3 and os_deposit > 1 and os_seam > 1:
                        cell_color = "872bff"
                    elif zin == 3 and mining_rating >= 140:
                        cell_color = "1dde26"
                    elif zin == 3 and mining_rating < 140:
                        cell_color = "f54242"
                    elif zin == 2:
                        cell_color = "ff9412"
                    elif intro_complete == 1:
                        cell_color = "000000"
                    elif intro_complete == 0:
                        cell_color = "b8b8b8"

                    cell.font = Font(color=cell_color, bold=cell_bold, underline=cell_underline)

        wb.save(EXCEL_PATH)
        conn.close()
    
    def update_realms_inventory(self, addon_data=False):
        """Get inventory info for each realm"""

        # Get auctions from MyAH_data_parser auctions DB
        auctions = {}
        conn = sqlite3.connect(AUCTIONS)
        c = conn.cursor()
        for realm in self.realms.values():
            auctions[realm.name] = {}
            c.execute("""SELECT item_id, quantity, stack_size FROM auction_chunks WHERE (
                    realm=? AND owner LIKE ?)""", (realm.name, realm.seller_name + '%'))
            for chunk in c.fetchall():
                auctions[realm.name][chunk[0]] = chunk[1] * chunk[2]
        conn.close()
        
        # Get inventory data from Multiboxer_Data (multiple chars on multiple accounts)
        inventory = {}
        for acc_data in self.accounts.values():
            for realm_name, realm_data in acc_data.realms.items():
                if realm_data['inventoryData']:
                    inventory[realm_name] = inventory.get(realm_name, {})
                    for char_inventory in realm_data['inventoryData'].values():
                        for item_id, quantity in char_inventory.items():
                            inventory[realm_name][item_id] = inventory[realm_name].get(item_id, 0) + quantity

        # Get auction data from Multiboxer_data (in case BLIZZ API is down)
        addon_auctions = {}
        for acc_data in self.accounts.values():
            for realm_name, realm_data in acc_data.realms.items():
                if realm_data['auctionData']:
                    addon_auctions[realm_name] = addon_auctions.get(realm_name, {})
                    for char_auctions in realm_data['auctionData'].values():
                        for item_id, quantity in char_auctions.items():
                            addon_auctions[realm_name][item_id] = addon_auctions[realm_name].get(item_id, 0) + quantity
        
        # use game auctions data if api is down
        auctions = addon_auctions if addon_data else auctions

        # Update RealmData objects with inventory and auction info
        for realm in self.realms.values():
            realm.update_inventory(auctions.get(realm.name, {}), inventory.get(realm.name, {}))

    def update_bankers(self):
        """doc"""
        for banker in self.bankers.values():
            accounting_data = self.accounts[banker.account].accounting
            bank_gold = accounting_data.get(banker.realm, {}).get(banker.name, {}).get('money', {}).get('guild', None)
            banker.update_info(bank_gold)

    def bankers_excel(self):
        conn = sqlite3.connect(CHAR_DB)
        c = conn.cursor()

        wb = load_workbook(EXCEL_PATH, read_only=False, keep_vba=True)
        ws = wb["Bankers"]

        for account in range(10):
            c.execute("""SELECT name, realm FROM char_db
                      WHERE role=? AND account=? AND type LIKE ?
                      ORDER BY id ASC""",
                      ('banker', account, 'gbank%'))
            rows = c.fetchall()
            for position, row in enumerate(rows):
                banker = self.bankers.get('-'.join((row[0], row[1])), None)
                cell = ws.cell(row=account+1, column=position+1)
                cell.value = banker.name
                if banker.bank_gold == 9_999_999:
                    cell.font = Font(color="fc0303", bold=True)
                elif not banker.bank_gold:
                    cell.font = Font(color="03befc", bold=False)

        conn.close()
        wb.save(EXCEL_PATH)


def pandas_inventory(wd, item_id):
    inventory_list = []
    for realm, realm_data in wd.realms.items():
        item_data = realm_data.inventory.get(item_id, {})
        bags = item_data.get('bags', 0) // 200
        ah = item_data.get('ah', 0) // 200
        total = bags + ah
        inventory_list.append((realm, total, ah, bags))

    sorted_list = sorted(inventory_list, key=lambda x: x[1], reverse=True)
    df = pd.DataFrame(sorted_list, columns=['realm', 'total', 'ah', 'bags'])
    print(df)


if __name__ == '__main__':
    wd = WowData()

    # wd.update_farmers()
    # wd.write_farmers_db()
    # wd.write_excel_table()

    wd.update_realms_inventory(addon_data=True)
    pandas_inventory(wd, 168487)
    
    # wd.update_bankers()
    # wd.bankers_excel()

    # total_g = 0
    # for banker in wd.bankers.values():
    #     total_g = total_g + (banker.bank_gold or 0)
    # print(total_g)
    
    input('\nPress any key to exit')